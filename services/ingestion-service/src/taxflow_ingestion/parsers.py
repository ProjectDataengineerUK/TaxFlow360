import csv
import io
from collections.abc import Iterator
from pathlib import Path
from typing import Any, Protocol
from xml.etree import ElementTree


class Parser(Protocol):
    def parse(self, content: bytes) -> Iterator[dict[str, Any]]: ...


class CsvParser:
    def parse(self, content: bytes) -> Iterator[dict[str, Any]]:
        stream = io.StringIO(content.decode("utf-8-sig"), newline="")
        yield from csv.DictReader(stream)


class XmlParser:
    def parse(self, content: bytes) -> Iterator[dict[str, Any]]:
        root = ElementTree.fromstring(content)
        records = list(root.findall(".//transaction"))
        if root.tag == "transaction":
            records.insert(0, root)
        for record in records:
            yield {child.tag: child.text or "" for child in record}


class XlsxParser:
    def parse(self, content: bytes) -> Iterator[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("XLSX support requires the openpyxl dependency") from exc
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            first_row = next(rows)
        except StopIteration as exc:
            raise ValueError("XLSX worksheet must include a header row") from exc
        headers = [str(value).strip() for value in first_row]
        if not all(headers):
            raise ValueError("XLSX headers must not be empty")
        for row in rows:
            yield dict(zip(headers, row, strict=False))


PARSERS: dict[str, Parser] = {".csv": CsvParser(), ".xml": XmlParser(), ".xlsx": XlsxParser()}


def parser_for(filename: str) -> Parser:
    extension = Path(filename).suffix.lower()
    try:
        return PARSERS[extension]
    except KeyError as exc:
        raise ValueError(f"unsupported file type: {extension or 'none'}") from exc
